# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from datetime import date
import time
from init import init_path_report


def create_report(arg2, arg3, arg4, arg5, arg6, arg7, arg9, arg10, arg11, arg12, arg13, arg14, arg15, arg16, arg17, arg18, arg20, arg21, arg22, arg23, arg24, arg25):
    # Load in the workbook
    #wb = load_workbook('C:\\Users\\Admin\\PycharmProjects\\AVA\\report.xlsx')
    wb = load_workbook(init_path_report())
    # Get sheet names
    #print(wb.get_sheet_names())

    # Get a sheet by name
    sheet = wb.get_sheet_by_name('Лист1')

    # Print the sheet title
    #print(sheet.title)

    # Retrieve the value of a certain cell

    sheet['B2'].value = arg2
    sheet['B3'].value = arg3
    sheet['B4'].value = arg4
    sheet['B5'].value = arg5
    sheet['B6'].value = arg6
    sheet['B7'].value = arg7
    sheet['B9'].value = arg9
    sheet['B10'].value = arg10
    sheet['B11'].value = arg11
    sheet['B12'].value = arg12
    sheet['B13'].value = arg13
    sheet['B14'].value = arg14
    sheet['B15'].value = arg15
    sheet['B16'].value = arg16
    sheet['B17'].value = arg17
    sheet['B18'].value = arg18


    sheet['D5'].value = arg20
    sheet['D7'].value = arg21
    sheet['D8'].value = arg22
    sheet['D9'].value = arg23
    sheet['D12'].value = arg24
    sheet['D18'].value = arg25
    wb.save(f"C:\\Users\\Admin\\PycharmProjects\\AVA\\report\\{date.today()}_{time.time()}_{arg25}.xlsx")


















"""import datetime
from datetime import date
from datetime import time
import time
from Parse_word import name_act #передаем данные из акта
import json
import os"""

"""def create_report():
    directory_folder = rf"C:\\Users\\Admin\\PycharmProjects\\AVA\\report\\{date.today()}_{time.time()}_{name_act}.json"
    folder_path = os.path.dirname(directory_folder) # Путь к папке с файлом

    if not os.path.exists(folder_path): #Если пути не существует создаем его
        os.makedirs(folder_path)

    with open(directory_folder, 'w') as file: # Открываем фаил и пишем
        file.write("этот текст создан автоматически")

    data = {}
    data['act'] = []
    data['jira'] = []
    data['act'].append({
        'name': name_act,
        'time': 'pythonist.ru',
        'raid': 'Nebraska'
    })
    data['jira'].append({
        'name': 'Larry',
        'time': 'pythonist.ru'
    })

    with open(f'C:\\Users\\Admin\\PycharmProjects\\AVA\\report\\file_{date.today()}_{time.time()}_{name_act}.json', 'w', "utf-8") as outfile:
        json.dump(data, outfile)"""

